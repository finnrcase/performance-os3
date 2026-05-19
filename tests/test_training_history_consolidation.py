import os
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app
from src import training as training_module
from tests.auth_helpers import configure_test_auth


def _row(days_ago: int, exercise: str, muscle_group: str, workout_id: str) -> dict:
    date = (pd.Timestamp.today().normalize() - pd.Timedelta(days=days_ago)).date().isoformat()
    return {
        **{column: "" for column in training_module.TRAINING_COLUMNS},
        "workout_id": workout_id,
        "date": date,
        "workout_type": "Strength",
        "muscle_group": muscle_group,
        "exercise": exercise,
        "set_number": 1,
        "sets": 3,
        "reps": 5,
        "weight": 225,
        "rpe": 8,
        "duration_minutes": 60,
        "source": "hevy",
        "external_id": f"{workout_id}:{exercise}:1",
        "hevy_workout_id": workout_id,
    }


class TrainingHistoryConsolidationTest(unittest.TestCase):
    def test_consolidates_old_rows_without_deleting_raw_data(self):
        rows = [
            _row(220, "Bench Press", "Chest", "old-1"),
            _row(210, "Bench Press", "Chest", "old-2"),
            _row(20, "Squat", "Legs", "recent-1"),
        ]
        source = pd.DataFrame(rows, columns=training_module.TRAINING_COLUMNS)

        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, {"DATABASE_URL": "", "TRAINING_RAW_WINDOW_DAYS": "180"}), patch.object(training_module, "WEEKLY_TRAINING_SUMMARY_PATH", Path(temp_dir) / "weekly.csv"), patch.object(training_module, "MONTHLY_TRAINING_SUMMARY_PATH", Path(temp_dir) / "monthly.csv"), patch.object(training_module, "EXERCISE_PR_HISTORY_PATH", Path(temp_dir) / "prs.csv"), patch.object(training_module, "MUSCLE_GROUP_VOLUME_HISTORY_PATH", Path(temp_dir) / "muscle.csv"):
            result = training_module.consolidate_old_training_history(training_df=source, cutoff_days=180)
            weekly = training_module.load_weekly_training_summary()
            monthly = training_module.load_monthly_training_summary()
            prs = training_module.load_exercise_pr_history()
            muscle = training_module.load_muscle_group_volume_history()

        self.assertEqual(result["raw_rows_total"], 3)
        self.assertEqual(result["raw_rows_summarized"], 2)
        self.assertEqual(result["raw_rows_deleted"], 0)
        self.assertFalse(weekly.empty)
        self.assertFalse(monthly.empty)
        self.assertFalse(prs.empty)
        self.assertFalse(muscle.empty)
        self.assertEqual(len(source), 3)

    def test_training_history_endpoint_caps_to_raw_window(self):
        client = TestClient(app)
        configure_test_auth(client)
        recent = pd.DataFrame([_row(10, "Squat", "Legs", "recent-1")], columns=training_module.TRAINING_COLUMNS)

        with patch.dict(os.environ, {"TRAINING_RAW_WINDOW_DAYS": "180"}), patch("backend.routes.training.load_live_training_log", return_value=recent) as load_live:
            response = client.get("/api/training/history?limit=25&days=3650")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["days"], 180)
        self.assertEqual(data["raw_window_days"], 180)
        load_live.assert_called_once()
        self.assertEqual(load_live.call_args.kwargs["days"], 180)

    def test_summary_status_reports_raw_window_counts_and_coaching_contract(self):
        client = TestClient(app)
        configure_test_auth(client)
        recent = pd.DataFrame([_row(10, "Squat", "Legs", "recent-1")], columns=training_module.TRAINING_COLUMNS)
        weekly = pd.DataFrame([{"period_start": "2026-01-05"}])
        monthly = pd.DataFrame([{"period_start": "2026-01-01"}])
        prs = pd.DataFrame([{"exercise": "Bench Press"}])
        muscle = pd.DataFrame([{"muscle_group": "Chest"}])

        with patch.dict(os.environ, {"TRAINING_RAW_WINDOW_DAYS": "180"}), patch("backend.routes.training.count_dataframe_rows", return_value=10), patch("backend.routes.training.load_live_training_log", return_value=recent), patch("backend.routes.training.load_weekly_training_summary", return_value=weekly), patch("backend.routes.training.load_monthly_training_summary", return_value=monthly), patch("backend.routes.training.load_exercise_pr_history", return_value=prs), patch("backend.routes.training.load_muscle_group_volume_history", return_value=muscle), patch("backend.routes.training.load_training_summary_state", return_value={"last_rebuilt_at": "2026-05-18T12:00:00+00:00"}):
            response = client.get("/api/training/summary/status")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["total_raw_rows"], 10)
        self.assertEqual(data["recent_raw_rows"], 1)
        self.assertEqual(data["older_raw_rows"], 9)
        self.assertIn("raw_features_preserved", data["coaching_contract"])

    def test_hevy_raw_export_returns_excel_workbook(self):
        client = TestClient(app)
        configure_test_auth(client)
        raw = pd.DataFrame([_row(10, "Bench Press", "Chest", "recent-1")], columns=training_module.TRAINING_COLUMNS)
        raw.loc[0, "notes"] = "Imported from Hevy | workout_title=Push Day | weight_unit=lb"
        weekly = pd.DataFrame([{"period_start": "2026-01-05", "workout_count": 3}])
        prs = pd.DataFrame([{"exercise": "Bench Press", "estimated_1rm": 275}])

        with patch("backend.routes.training.load_training_log", return_value=raw), patch("backend.routes.training.load_weekly_training_summary", return_value=weekly), patch("backend.routes.training.load_exercise_pr_history", return_value=prs), patch("backend.routes.training._training_summary_status", return_value={"raw_window_days": 180, "total_raw_rows": 1, "recent_raw_rows": 1, "older_raw_rows": 0}):
            response = client.get("/api/training/export/hevy-raw")

        self.assertEqual(response.status_code, 200)
        self.assertIn("hevy_raw_export_", response.headers["content-disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertIn("raw_sets", workbook.sheetnames)
        self.assertIn("workouts_summary", workbook.sheetnames)
        self.assertIn("exercise_prs", workbook.sheetnames)
        self.assertIn("weekly_summary", workbook.sheetnames)
        self.assertIn("metadata", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
